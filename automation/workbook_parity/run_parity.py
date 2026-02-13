#!/usr/bin/env python3
"""
Workbook ↔ Lake parity harness for "Investor Questions - GWI Business.xlsx".

This is a *read-only* validator:
- Extract expected metrics from the workbook.
- Compute the same metrics from governed lake views (Athena).
- Compare like-for-like definitions and numbers.
- Write local artifacts (status + comparisons + extracted workbook slices).
"""

from __future__ import annotations

import csv
import datetime as dt
import json
import os
import re
import sys
import time
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

import boto3
from openpyxl import load_workbook


DEFAULT_WORKBOOK = "/Users/patch/Downloads/batch 9 vetro (act design)/Investor Questions - GWI Business (2).xlsx"
DEFAULT_OUTDIR = "/Users/patch/lake_deploy/deliverables/2026-02-13_workbook_parity"

AWS_REGION = os.environ.get("AWS_REGION", "us-east-2")
ATHENA_WORKGROUP = os.environ.get("ATHENA_WORKGROUP", "primary")
ATHENA_OUTPUT = os.environ.get(
    "ATHENA_OUTPUT_LOCATION",
    "s3://gwi-raw-us-east-2-pc/athena-results/orchestration/",
)


def norm(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"[^a-z0-9]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def json_dump(path: str, obj: Any) -> None:
    def _default(x: Any) -> str:
        if isinstance(x, (dt.date, dt.datetime)):
            return x.date().isoformat() if isinstance(x, dt.datetime) else x.isoformat()
        return str(x)

    with open(path, "w", encoding="utf-8") as fh:
        json.dump(obj, fh, indent=2, sort_keys=True, default=_default)


def csv_dump(path: str, rows: List[Dict[str, Any]]) -> None:
    if not rows:
        with open(path, "w", encoding="utf-8") as fh:
            fh.write("")
        return
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)


@dataclass
class AthenaResult:
    qid: str
    columns: List[str]
    rows: List[List[str]]


class Athena:
    def __init__(self) -> None:
        self.client = boto3.client("athena", region_name=AWS_REGION)

    def query(self, sql: str, database: str) -> AthenaResult:
        qid = self.client.start_query_execution(
            QueryString=sql,
            QueryExecutionContext={"Database": database},
            WorkGroup=ATHENA_WORKGROUP,
            ResultConfiguration={"OutputLocation": ATHENA_OUTPUT},
        )["QueryExecutionId"]

        while True:
            st = self.client.get_query_execution(QueryExecutionId=qid)["QueryExecution"]["Status"]["State"]
            if st in {"SUCCEEDED", "FAILED", "CANCELLED"}:
                break
            time.sleep(0.75)

        if st != "SUCCEEDED":
            status = self.client.get_query_execution(QueryExecutionId=qid)["QueryExecution"]["Status"]
            raise RuntimeError(f"Athena query failed ({st}) qid={qid}: {status.get('StateChangeReason')}")

        rs = self.client.get_query_results(QueryExecutionId=qid, MaxResults=1000)["ResultSet"]
        cols = [c["Name"] for c in rs["ResultSetMetadata"]["ColumnInfo"]]
        rows: List[List[str]] = []
        for row in rs["Rows"][1:]:
            rows.append([d.get("VarCharValue") if d else None for d in row.get("Data", [])])
        return AthenaResult(qid=qid, columns=cols, rows=rows)


def extract_customer_mix(wb) -> Dict[str, Any]:
    ws = wb["Customer Mix"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    idx = {str(h).strip(): i for i, h in enumerate(headers) if h not in (None, "")}
    required = ["Network", "Network Type", "Customer Type", "Access Type", "Passings", "Subscriptions", "ARPU"]
    for r in required:
        if r not in idx:
            raise RuntimeError(f"Customer Mix missing required column: {r}")

    records: List[Dict[str, Any]] = []
    for row in range(2, ws.max_row + 1):
        network = ws.cell(row, idx["Network"] + 1).value
        if network in (None, ""):
            continue
        records.append(
            {
                "network": str(network).strip(),
                "network_type": str(ws.cell(row, idx["Network Type"] + 1).value or "").strip(),
                "customer_type": str(ws.cell(row, idx["Customer Type"] + 1).value or "").strip(),
                "access_type": str(ws.cell(row, idx["Access Type"] + 1).value or "").strip(),
                "passings": float(ws.cell(row, idx["Passings"] + 1).value or 0),
                "subscriptions": float(ws.cell(row, idx["Subscriptions"] + 1).value or 0),
                "arpu_cell": ws.cell(row, idx["ARPU"] + 1).value,
            }
        )

    combos: Dict[Tuple[str, str, str], Dict[str, float]] = {}
    for r in records:
        key = (r["network_type"], r["customer_type"], r["access_type"])
        slot = combos.setdefault(key, {"passings": 0.0, "subscriptions": 0.0})
        slot["passings"] += float(r["passings"])
        slot["subscriptions"] += float(r["subscriptions"])

    combo_rows = [
        {"network_type": k[0], "customer_type": k[1], "access_type": k[2], **v}
        for k, v in sorted(combos.items(), key=lambda kv: kv[0])
    ]
    return {"records": records, "combos": combo_rows}


def find_revenue_mix_header_dates(ws) -> List[dt.date]:
    dates: List[dt.date] = []
    for col in range(1, ws.max_column + 1):
        v = ws.cell(1, col).value
        if isinstance(v, dt.datetime):
            dates.append(v.date())
        elif isinstance(v, dt.date):
            dates.append(v)
    return dates


def extract_revenue_mix(wb, target_as_of_date: Optional[dt.date] = None) -> Dict[str, Any]:
    ws = wb["Revenue Mix"]
    header_dates = find_revenue_mix_header_dates(ws)
    if not header_dates:
        raise RuntimeError("Revenue Mix: missing date headers in row 1")
    as_of_date = target_as_of_date or max(header_dates)

    asof_col = None
    for col in range(1, ws.max_column + 1):
        v = ws.cell(1, col).value
        if isinstance(v, dt.datetime) and v.date() == as_of_date:
            asof_col = col
            break
        if isinstance(v, dt.date) and v == as_of_date:
            asof_col = col
            break
    if not asof_col:
        raise RuntimeError(f"Revenue Mix: could not find column for as_of_date={as_of_date} (available={sorted(set(header_dates))[:12]}...)")

    def read_summary_block(row_start: int, row_end: int) -> List[Dict[str, Any]]:
        out: List[Dict[str, Any]] = []
        for r in range(row_start, row_end + 1):
            seg = ws.cell(r, 1).value
            access = ws.cell(r, 2).value
            val = ws.cell(r, asof_col).value
            if not isinstance(seg, str) or not seg.strip():
                continue
            if val is None:
                continue
            try:
                num = float(val)
            except Exception:
                continue
            out.append(
                {
                    "segment": seg.strip(),
                    "access_type": str(access or "").strip() or None,
                    "value": num,
                    "row": r,
                }
            )
        return out

    # These ranges are stable for the workbook format we are enforcing parity against.
    revenue_summary = read_summary_block(464, 472)
    revenue_total_excl_dvfiber = read_summary_block(474, 474)
    plat_summary = read_summary_block(476, 484)
    plat_total_excl_dvfiber = read_summary_block(486, 486)
    arpu_summary = read_summary_block(488, 496)
    arpu_total_excl_dvfiber = read_summary_block(498, 498)

    return {
        "as_of_date": as_of_date.isoformat(),
        "asof_col": asof_col,
        "revenue_summary": revenue_summary,
        "revenue_total_excluding_dvfiber": revenue_total_excl_dvfiber,
        "plat_id_count_summary": plat_summary,
        "plat_id_count_total_excluding_dvfiber": plat_total_excl_dvfiber,
        "arpu_summary": arpu_summary,
        "arpu_total_excluding_dvfiber": arpu_total_excl_dvfiber,
    }


def compare_exact(metric: str, expected: float, actual: float) -> Dict[str, Any]:
    return {"metric": metric, "expected": expected, "actual": actual, "delta": actual - expected, "ok": expected == actual}


def compare_tol(metric: str, expected: float, actual: float, tol: float) -> Dict[str, Any]:
    return {
        "metric": metric,
        "expected": expected,
        "actual": actual,
        "delta": actual - expected,
        "tol": tol,
        "ok": abs(actual - expected) <= tol,
    }


def main() -> int:
    workbook_path = os.environ.get("WORKBOOK_PATH", DEFAULT_WORKBOOK)
    outdir = os.environ.get("OUTDIR", DEFAULT_OUTDIR)
    os.makedirs(outdir, exist_ok=True)

    wb = load_workbook(workbook_path, data_only=False)
    cm = extract_customer_mix(wb)

    # Anchor Revenue Mix extraction to the lake's "latest as_of_date" so we compare the same month.
    ath = Athena()
    asof_sql = "SELECT CAST(MAX(as_of_date) AS varchar) AS as_of_date FROM curated_core.v_investor_revenue_mix_latest;"
    asof_res = ath.query(asof_sql, database="curated_core")
    lake_asof = None
    if asof_res.rows and asof_res.rows[0] and asof_res.rows[0][0]:
        lake_asof = dt.date.fromisoformat(str(asof_res.rows[0][0])[:10])
    rm = extract_revenue_mix(wb, target_as_of_date=lake_asof)

    json_dump(os.path.join(outdir, "workbook_customer_mix.json"), cm)
    json_dump(os.path.join(outdir, "workbook_revenue_mix.json"), rm)

    qids: Dict[str, str] = {}
    qids["lake_revenue_mix_max_as_of_date"] = asof_res.qid

    cm_sql = """
SELECT
  network_type,
  customer_type,
  CASE WHEN LOWER(COALESCE(network_type, '')) = 'clec' THEN 'Copper' ELSE 'Fiber' END AS access_type,
  SUM(COALESCE(passings, 0)) AS passings,
  SUM(COALESCE(subscriptions, 0)) AS subscriptions
FROM curated_core.v_network_health
WHERE dt = (SELECT MAX(dt) FROM curated_core.v_network_health)
  AND network IS NOT NULL
  AND TRIM(network) <> ''
  AND network <> 'Unmapped'
GROUP BY 1,2,3
ORDER BY 1,2,3;
""".strip()
    cm_res = ath.query(cm_sql, database="curated_core")
    qids["lake_customer_mix_combos"] = cm_res.qid

    lake_cm = []
    for r in cm_res.rows:
        lake_cm.append(
            {
                "network_type": r[0],
                "customer_type": r[1],
                "access_type": r[2],
                "passings": float(r[3] or 0),
                "subscriptions": float(r[4] or 0),
            }
        )

    rm_sql = """
SELECT
  network_type,
  MAX(as_of_date) AS as_of_date,
  SUM(COALESCE(plat_id_count, 0)) AS plat_id_count,
  SUM(COALESCE(revenue, 0)) AS revenue
FROM curated_core.v_investor_revenue_mix_latest
WHERE network IS NOT NULL
  AND TRIM(network) <> ''
  AND network <> 'Total'
GROUP BY 1
ORDER BY 1;
""".strip()
    rm_res = ath.query(rm_sql, database="curated_core")
    qids["lake_revenue_mix_by_type"] = rm_res.qid

    lake_rm = []
    for r in rm_res.rows:
        lake_rm.append(
            {
                "network_type": r[0],
                "as_of_date": r[1],
                "plat_id_count": float(r[2] or 0),
                "revenue": float(r[3] or 0),
            }
        )

    json_dump(os.path.join(outdir, "athena_qids.json"), qids)
    json_dump(os.path.join(outdir, "lake_customer_mix_combos.json"), lake_cm)
    json_dump(os.path.join(outdir, "lake_revenue_mix_by_type.json"), lake_rm)

    # Workbook includes a "Resold" copper rollup row (distinct from "Resold; Copper").
    # Compute the same rollup from the lake for parity checks.
    copper_rollup_sql = """
SELECT
  MAX(as_of_date) AS as_of_date,
  SUM(COALESCE(plat_id_count, 0)) AS plat_id_count,
  SUM(COALESCE(revenue, 0)) AS revenue
FROM curated_core.v_investor_revenue_mix_latest
WHERE network IS NOT NULL
  AND TRIM(network) <> ''
  AND network <> 'Total'
  AND LOWER(COALESCE(network_type, '')) LIKE '%copper%';
""".strip()
    copper_rollup_res = ath.query(copper_rollup_sql, database="curated_core")
    qids["lake_revenue_mix_copper_rollup"] = copper_rollup_res.qid
    lake_copper_rollup = {
        "as_of_date": (copper_rollup_res.rows[0][0] if copper_rollup_res.rows else None),
        "plat_id_count": float(copper_rollup_res.rows[0][1] or 0) if copper_rollup_res.rows else 0.0,
        "revenue": float(copper_rollup_res.rows[0][2] or 0) if copper_rollup_res.rows else 0.0,
    }
    json_dump(os.path.join(outdir, "lake_revenue_mix_copper_rollup.json"), lake_copper_rollup)

    comparisons: List[Dict[str, Any]] = []
    failures: List[Dict[str, Any]] = []

    exp_cm = {(norm(r["network_type"]), norm(r["customer_type"]), norm(r["access_type"])): r for r in cm["combos"]}
    act_cm = {(norm(r["network_type"]), norm(r["customer_type"]), norm(r["access_type"])): r for r in lake_cm}

    for k, exp in exp_cm.items():
        act = act_cm.get(k)
        if not act:
            failures.append({"type": "missing_combo", "combo": k})
            continue
        comparisons.append(compare_exact(f"customer_mix.passings.{k}", float(exp["passings"]), float(act["passings"])))
        comparisons.append(compare_exact(f"customer_mix.subscriptions.{k}", float(exp["subscriptions"]), float(act["subscriptions"])))

    # Revenue mix: compare workbook summary blocks to lake by network_type (exact).
    lake_rm_by_type = {norm(r["network_type"]): r for r in lake_rm}

    for row in rm["revenue_summary"]:
        seg = row["segment"]
        if norm(seg) == "resold":
            # rollup: all copper revenue (matches workbook "Resold" / Copper)
            comparisons.append(compare_tol("revenue_mix.revenue.Resold (Copper rollup)", float(row["value"]), float(lake_copper_rollup["revenue"]), tol=0.01))
            continue
        act = lake_rm_by_type.get(norm(seg))
        if not act:
            continue
        comparisons.append(compare_tol(f"revenue_mix.revenue.{seg}", float(row["value"]), float(act["revenue"]), tol=0.01))

    for row in rm["plat_id_count_summary"]:
        seg = row["segment"]
        if norm(seg) == "resold":
            comparisons.append(compare_tol("revenue_mix.plat_id_count.Resold (Copper rollup)", float(row["value"]), float(lake_copper_rollup["plat_id_count"]), tol=0.01))
            continue
        act = lake_rm_by_type.get(norm(seg))
        if not act:
            continue
        comparisons.append(compare_tol(f"revenue_mix.plat_id_count.{seg}", float(row["value"]), float(act["plat_id_count"]), tol=0.01))

    for row in rm["arpu_summary"]:
        seg = row["segment"]
        if norm(seg) == "resold":
            denom = float(lake_copper_rollup["plat_id_count"]) or 0.0
            actual_arpu = (float(lake_copper_rollup["revenue"]) / denom) if denom > 0 else 0.0
            comparisons.append(compare_tol("revenue_mix.arpu.Resold (Copper rollup)", float(row["value"]), actual_arpu, tol=0.5))
            continue
        act = lake_rm_by_type.get(norm(seg))
        if not act:
            continue
        denom = float(act["plat_id_count"]) or 0.0
        actual_arpu = (float(act["revenue"]) / denom) if denom > 0 else 0.0
        comparisons.append(compare_tol(f"revenue_mix.arpu.{seg}", float(row["value"]), actual_arpu, tol=0.5))

    ok = not failures and all(c.get("ok") for c in comparisons)
    status = {
        "run_at_utc": dt.datetime.utcnow().isoformat() + "Z",
        "workbook_path": workbook_path,
        "workbook_revenue_mix_as_of_date": rm["as_of_date"],
        "result": "PASS" if ok else "FAIL",
        "comparisons_total": len(comparisons),
        "comparisons_failed": len([c for c in comparisons if not c.get("ok")]),
        "failures": failures,
        "athena_qids": qids,
    }
    json_dump(os.path.join(outdir, "status.json"), status)
    json_dump(os.path.join(outdir, "comparisons.json"), comparisons)

    # Compact MD for humans
    report = [
        "# Workbook ↔ Lake Parity",
        f"- Run at (UTC): `{status['run_at_utc']}`",
        f"- Workbook: `{workbook_path}`",
        f"- Revenue Mix as_of_date (workbook header selection): `{rm['as_of_date']}`",
        f"- Result: **{status['result']}**",
        "",
        "## Athena QIDs",
    ]
    for k, v in qids.items():
        report.append(f"- `{k}`: `{v}`")
    report.append("")
    report.append("## Failures")
    if not failures:
        report.append("- (none)")
    else:
        for f in failures[:100]:
            report.append(f"- {f}")
    report.append("")
    report.append("## Comparisons (failed only)")
    failed = [c for c in comparisons if not c.get("ok")]
    if not failed:
        report.append("- (none)")
    else:
        for c in failed[:200]:
            report.append(f"- `{c['metric']}` expected={c['expected']} actual={c['actual']} delta={c['delta']}")
    with open(os.path.join(outdir, "report.md"), "w", encoding="utf-8") as fh:
        fh.write("\n".join(report) + "\n")

    return 0 if ok else 2


if __name__ == "__main__":
    raise SystemExit(main())
